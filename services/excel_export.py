from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def exportar_ventas_xlsx(ruta, headers, filas):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    ws.append(headers)
    encabezado_fill = PatternFill(start_color="5CB85C", end_color="5CB85C", fill_type="solid")
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = encabezado_fill
        celda.alignment = Alignment(horizontal="center")

    for fila in filas:
        ws.append(fila)

    ancho_base = 14
    for col_idx in range(1, len(headers) + 1):
        col_letra = get_column_letter(col_idx)
        prueba = max(len(str(headers[col_idx - 1])), 12)
        for fila in filas:
            if col_idx - 1 < len(fila):
                prueba = max(prueba, len(str(fila[col_idx - 1])))
        ws.column_dimensions[col_letra].width = min(prueba + 2, 40)

    wb.save(ruta)